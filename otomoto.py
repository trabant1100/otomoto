import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from io import BytesIO
from PIL import Image as PILImage
from datetime import date


output = date.today().strftime("%d.%m.%Y") + ".xlsx"
results = []
URL = "https://www.otomoto.pl/osobowe/ford/explorer/od-2020?search%5Bfilter_float_year%3Ato%5D=2020&search%5Badvanced_search_expanded%5D=false"
page = requests.get(URL)

soap = BeautifulSoup(page.content, "html.parser")
articles = soap.find_all("article")

for article in articles:
	result = {}
	if article.section:
		for div in article.section.children:
			img = div.find("img", recursive=False)
			h1 = div.find("h1", recursive=False)
			p = div.find("p", recursive=False)
			dls = div.find_all("dl", recursive=False)
			if img:
				result["thumbnail"] = img["src"]
			if h1:
				result["title"] = h1.a.string
				result["url"] = h1.a["href"]
			if p:
				result["description"] = p.string
			if dls:
				result["mileage"] = "".join(dls[0].dd.strings)
				result["location"] = dls[1].dd.p.string
			if div.h3:
				result["price"] = div.h3.string
		if result.get("thumbnail") is None:
			img = article.section.div.find("img")
			result["thumbnail"] = img["src"]

	if result:
		results.append(result)
		
for idx, item in enumerate(results):
	print(idx, item.get("thumbnail"), item.get("description"), item.get("title"), 
		item.get("url"), item.get("description"), item.get("location"), item.get("mileage"), item.get("price"))	


def calculate_resized_dimensions(original_width, original_height, max_width, max_height):
    ratio = min(max_width / original_width, max_height / original_height)
    return int(original_width * ratio), int(original_height * ratio)

def resize_column_to_fit_text(ws, column_letter):
    """
    Resizes the specified column in the given worksheet to fit the text content.
    
    :param ws: The worksheet object.
    :param column_letter: The column letter (e.g., 'A', 'B') to resize.
    """
    max_length = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.column_letter == column_letter and cell.value:
                max_length = max(max_length, len(str(cell.value)))
    
    # Adding extra space for padding
    adjusted_width = max_length + 2

    # Set the column width
    ws.column_dimensions[column_letter].width = adjusted_width


# Create a workbook and a sheet
wb = Workbook()
ws = wb.active
ws.title = "otomoto"

# Add some data
header = ["thumbnail", "title", "url", "description", "location", "mileage", "price"]
ws.append(header)
for idx, item in enumerate(results):
	values = list(map(lambda h: item[h], header))
	ws.append(values)
	response = requests.get(item.get("thumbnail"))
	image_data = BytesIO(response.content)
	img_pil = PILImage.open(image_data)

	max_width = 400  # Maximum width
	max_height = 300  # Maximum height
	resized_width, resized_height = calculate_resized_dimensions(img_pil.width, img_pil.height, max_width, max_height)

	img_pil = img_pil.resize((resized_width, resized_height), PILImage.LANCZOS)
	img_bytes = BytesIO()
	img_pil.save(img_bytes, format='PNG')
	img_xl = XLImage(img_bytes)
	cell = ws.cell(row=(idx+2), column=1)
	ws.add_image(img_xl, cell.coordinate)
	ws.row_dimensions[idx+2].height = 250
	ws.column_dimensions["A"].width = 400 / 7

for idx in range(1, len(header)):
	resize_column_to_fit_text(ws, get_column_letter(idx + 1))


# Save the workbook to a file
wb.save(output)	