import glob
import os
import json
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from collections import defaultdict
import re
import requests
from io import BytesIO

with open("config.json", "r") as file:
    config = json.load(file)
    listing_config = config["listing"]
    report_config = config["report"]
    DIR = report_config["dir"]
    banned_urls = report_config["banned_urls"]
    crashed_urls = report_config["crashed_urls"]
    fav_urls = report_config["fav_urls"]
    dead_urls = report_config["dead_urls"]

def read_all_excel_files_to_dict(directory, extension):
    if not extension.startswith('.'):
        extension = '.' + extension

    pattern = os.path.join(directory, f'*{extension}')
    files = glob.glob(pattern)

    all_data = {}

    for file in files:
        # Check if the file matches the expected date format in the filename
        if re.match(r'\d{2}\.\d{2}\.\d{4}\.xlsx', os.path.basename(file)):
            try:
                wb = load_workbook(file)
                workbook_data = []
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    headers = [cell.value for cell in ws[1]]  # Assuming first row is the header
                    # Add a Date column based on the filename
                    headers.insert(0, 'Date')
                    date_value = datetime.strptime(os.path.splitext(os.path.basename(file))[0], '%d.%m.%Y').date()
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        row_data = {headers[i]: row[i-1] if i > 0 else date_value for i in range(len(headers))}
                        workbook_data.append(row_data)
                # Sort workbook_data by 'Date' column
                workbook_data.sort(key=lambda x: x['Date'])
                all_data[os.path.basename(file)] = workbook_data
            except Exception as e:
                print(f"Error opening {file}: {e}")

    return all_data

def group_by_thumbnail(all_workbooks_data, thumbnail_column):
    grouped_data = defaultdict(list)
    
    for workbook, data in all_workbooks_data.items():
        for row in data:
            thumbnail = row.get(thumbnail_column)
            if thumbnail:
                grouped_data[thumbnail].append((workbook, row))
    
    return grouped_data

def sanitize_sheet_title(title, max_length=30):
    # Replace invalid characters with underscores or remove them
    sanitized_title = re.sub(r'[\\/*?[\]:]', '_', title)
    # Ensure the title is no longer than the maximum allowed length
    if len(sanitized_title) > max_length:
        sanitized_title = sanitized_title[:max_length]
    return sanitized_title

def ordinal(n):
    suffix = {1: 'st', 2: 'nd', 3: 'rd'}
    return f"{n}{suffix.get(n if n not in {11, 12, 13} else n % 10, 'th')}"

def adjust_column_widths(ws):
    for col in ws.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2  # Adjust as needed
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = adjusted_width

def calculate_resized_dimensions(original_width, original_height, max_width, max_height):
    ratio = min(max_width / original_width, max_height / original_height)
    return int(original_width * ratio), int(original_height * ratio)

def insert_image_from_url(ws, url, row_index, image_width=320, image_height=240):
    try:
        response = requests.get(url)
        image_data = BytesIO(response.content)
        img_pil = PILImage.open(image_data)

        max_width = 400  # Maximum width
        max_height = 300  # Maximum height
        resized_width, resized_height = calculate_resized_dimensions(img_pil.width, img_pil.height, max_width, max_height)

        img_pil = img_pil.resize((resized_width, resized_height), PILImage.LANCZOS)
        img_bytes = BytesIO()
        img_pil.save(img_bytes, format='PNG')
        img_xl = XLImage(img_bytes)
        cell = ws.cell(row=2, column=1)
        ws.add_image(img_xl, cell.coordinate)
    except Exception as e:
        print(f"Error inserting image from {url}: {e}")

if __name__ == "__main__":
    directory = listing_config["dir"]
    extension = 'xlsx'
    thumbnail_column = 'thumbnail'  # Replace with your actual column name for thumbnails

    all_workbooks_data = read_all_excel_files_to_dict(directory, extension)

    grouped_data = group_by_thumbnail(all_workbooks_data, thumbnail_column)

    output_wb = Workbook()
    output_wb.remove(output_wb.active)  # Remove default sheet

    for idx, (thumbnail, entries) in enumerate(grouped_data.items(), start=1):
        ordinal_number = ordinal(idx)
        desc = entries[0][1]['description']
        url = entries[0][1]['url']
        if url in banned_urls:
            desc = f"ZZZ{desc}"
        sheet_title = sanitize_sheet_title(f"{desc}")
        ws = output_wb.create_sheet(title=f"{idx}")
        if url in banned_urls or url in crashed_urls:
            ws.sheet_state = 'hidden'
        if url in dead_urls:
            ws.sheet_properties.tabColor = 'D3D3D3'
        if url in fav_urls:
            ws.sheet_properties.tabColor = '00FF00'    
        if entries:
            headers = ['img', 'price'] + list(entries[0][1].keys())
            ws.append(headers)  # Add headers as the first row
            # Sort entries by 'Date' column before appending to worksheet
            entries.sort(key=lambda x: x[1]['Date'])
            for row_index, (workbook, row) in enumerate(entries, start=2):  # Start from row 2 (after headers)
                ws.append(['', row['price']] + list(row.values()))  # Append row data
                # Insert image from the first thumbnail column into a new column 'B'
                thumbnail_url = row.get(thumbnail_column)
                if thumbnail_url and row_index == 2:  # Only insert the first image
                    insert_image_from_url(ws, thumbnail_url, row_index)
            if row_index == 2:
                ws.sheet_properties.tabColor = 'FFA500'

        # Adjust column widths to fit content
        adjust_column_widths(ws)
        ws.column_dimensions["A"].width = 400 / 7
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["F"].width = 20
        ws['F2'].hyperlink = url

    output_file = DIR + "/" + date.today().strftime("%d.%m.%Y") + " report.xlsx"  
    output_wb.save(output_file)
    print(f"Data written to {output_file}")
